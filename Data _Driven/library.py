import openpyxl
from datetime import datetime

class Comman_Testcases:

    def __init__(self, filepath, sheetname):
        self.wk_book = openpyxl.load_workbook(filepath)
        self.sht = self.wk_book[sheetname]

    def fetch_row_count(self):
        rows = self.sht.max_row
        return rows

    def fetch_column_count(self):
        columns = self.sht.max_column
        return columns

    def fetch_keys_name(self):
        lst = []
        for i in range(1, self.sht.max_column + 1):
            col = self.sht.cell(row=1, column=i)
            lst.append(col.value)
        return lst

    def update_data(self, row_number, json_request, keylist):
        for i in range(1, self.sht.max_column + 1):
            cell = self.sht.cell(row=row_number, column=i)
            value = cell.value
            if isinstance(value, datetime):
                value = value.isoformat()  # Convert datetime to ISO format string
            json_request[keylist[i - 1]] = value
        return json_request
