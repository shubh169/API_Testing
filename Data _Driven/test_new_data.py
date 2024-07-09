import requests
import json
import openpyxl

def test_adding_students():
    url="https://thetestingworldapi.com/api/studentsDetails"
    with open('student_data.json','r') as data:
        student_data=data.read()

    json_request=json.loads(student_data)

# Load the workbook and select the sheet.
    wk_book=openpyxl.load_workbook('test_data.xlsx')
    sht=wk_book['Sheet1']
    rows=sht.max_row

# Iterate through the rows and print the cell values.
    for i in range(2,rows+1):
        cell_first_name=sht.cell(row=i,column=1)
        cell_mid_name = sht.cell(row=i, column=2)
        cell_last_name = sht.cell(row=i, column=3)
        cell_dob = sht.cell(row=i, column=4)

# Put excel data in json.
        json_request['first_name']=cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value


        response=requests.post(url,json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code==201
