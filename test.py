import openpyxl

# Load the workbook and select the sheet
wk_book = openpyxl.load_workbook('test_data.xlsx')
sht = wk_book['Sheet1']
rows = sht.max_row

# Iterate through the rows and print the cell values
for i in range(2, rows + 1):
    cell_first_name = sht.cell(row=i, column=1).value
    cell_mid_name = sht.cell(row=i, column=2).value
    cell_last_name = sht.cell(row=i, column=3).value
    cell_dob = sht.cell(row=i, column=4).value
    print(f"First Name: {cell_first_name}, Middle Name: {cell_mid_name}, Last Name: {cell_last_name}, DOB: {cell_dob}")
